import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================
# Warna
# ============================================================
MERAH       = "FFFF4444"
HIJAU       = "FF44BB44"
KUNING      = "FFFFC107"
BIRU_TUA    = "FF1565C0"
ABU         = "FFF5F5F5"
PUTIH       = "FFFFFFFF"
HEADER_BG   = "FF1A237E"
NEGATIF_BG  = "FFFFEBEE"
POSITIF_BG  = "FFE8F5E9"

def style_header(cell, bg_color=HEADER_BG, font_color="FFFFFFFF", bold=True, size=11):
    cell.font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(cell, bg_color=PUTIH, font_color="FF000000", bold=False, center=False):
    cell.font = Font(color=font_color, bold=bold, size=10, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True
    )

def thin_border():
    thin = Side(style="thin", color="FFCCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================
# Sheet 1 – Dashboard Ringkasan
# ============================================================
def buat_sheet_dashboard(wb, stats: dict, keyword: str):
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    # Judul
    ws.merge_cells("A1:B1")
    ws["A1"] = "📋 LAPORAN REKAP INSTAGRAM"
    ws["A1"].font = Font(bold=True, size=16, color=PUTIH, name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Keyword: {keyword}   |   Dibuat: {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws["A2"].font = Font(size=10, italic=True, color="FF555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Data stats
    items = [
        ("🔍 Keyword Target",       keyword,                                    BIRU_TUA),
        ("📄 Total Post Dikumpulkan", stats.get("total_posts", 0),              BIRU_TUA),
        ("💬 Total Komentar",        stats.get("total_comments", 0),            BIRU_TUA),
        ("🔴 Komentar Negatif",      stats.get("total_negative_comments", 0),   "FFCC0000"),
        ("🟢 Komentar Positif",
            stats.get("total_comments", 0) - stats.get("total_negative_comments", 0),
            "FF2E7D32"),
        ("📅 Tanggal Export",        datetime.now().strftime("%d-%m-%Y %H:%M"), "FF555555"),
    ]

    for i, (label, value, color) in enumerate(items, start=4):
        ws.row_dimensions[i].height = 28
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True, size=11, name="Calibri")
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=ABU)
        ws[f"A{i}"].alignment = Alignment(vertical="center")
        ws[f"A{i}"].border = thin_border()

        ws[f"B{i}"] = value
        ws[f"B{i}"].font = Font(bold=True, size=12, color=color, name="Calibri")
        ws[f"B{i}"].fill = PatternFill("solid", fgColor=PUTIH)
        ws[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{i}"].border = thin_border()

# ============================================================
# Sheet 2 – Semua Post
# ============================================================
def buat_sheet_posts(wb, posts: list):
    ws = wb.create_sheet("📄 Semua Post")

    headers = ["No", "URL Post", "Username", "Caption", "Likes", "Komentar", "Tanggal", "Keyword"]
    lebar   = [5,    45,         18,          50,        10,      12,          20,         20]

    for col, (h, w) in enumerate(zip(headers, lebar), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, post in enumerate(posts, 1):
        ws.row_dimensions[i + 1].height = 45
        bg = ABU if i % 2 == 0 else PUTIH
        row_data = [
            i,
            post.get("url", ""),
            post.get("username", ""),
            post.get("caption", "")[:200],
            post.get("likes", 0),
            post.get("comments", 0),
            post.get("tanggal", ""),
            post.get("keyword", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            style_cell(cell, bg_color=bg, center=(col in [1, 5, 6]))
            cell.border = thin_border()

# ============================================================
# Sheet 3 – Komentar Negatif
# ============================================================
def buat_sheet_komentar_negatif(wb, comments: list):
    ws = wb.create_sheet("🔴 Komentar Negatif")

    headers = ["No", "Post URL", "Username", "Komentar", "Skor Negatif", "Label"]
    lebar   = [5,    45,         18,          60,         15,              12]

    for col, (h, w) in enumerate(zip(headers, lebar), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell, bg_color="FFCC0000")
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, c in enumerate(comments, 1):
        ws.row_dimensions[i + 1].height = 50
        row_data = [
            i,
            c.get("post_url", ""),
            c.get("username", ""),
            c.get("text", ""),
            c.get("negative_score", 0),
            "🔴 NEGATIF",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            style_cell(cell, bg_color=NEGATIF_BG, center=(col in [1, 5, 6]))
            if col == 6:
                cell.font = Font(bold=True, color="FFCC0000", size=10)
            cell.border = thin_border()

# ============================================================
# Sheet 4 – Semua Komentar
# ============================================================
def buat_sheet_semua_komentar(wb, comments: list):
    ws = wb.create_sheet("💬 Semua Komentar")

    headers = ["No", "Post URL", "Username", "Komentar", "Status"]
    lebar   = [5,    45,         18,          60,         12]

    for col, (h, w) in enumerate(zip(headers, lebar), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, c in enumerate(comments, 1):
        ws.row_dimensions[i + 1].height = 50
        is_neg = c.get("is_negative", 0) == 1
        bg = NEGATIF_BG if is_neg else POSITIF_BG
        label = "🔴 NEGATIF" if is_neg else "🟢 POSITIF"

        row_data = [
            i,
            c.get("post_url", ""),
            c.get("username", ""),
            c.get("text", ""),
            label,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            style_cell(cell, bg_color=bg, center=(col in [1, 5]))
            if col == 5:
                color = "FFCC0000" if is_neg else "FF2E7D32"
                cell.font = Font(bold=True, color=color, size=10)
            cell.border = thin_border()

# ============================================================
# MAIN: Generate Excel
# ============================================================
def export_ke_excel(
    posts: list,
    semua_komentar: list,
    komentar_negatif: list,
    stats: dict,
    keyword: str,
    filename: str = None
) -> str:
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rekap_instagram_{keyword}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    # Hapus sheet default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buat_sheet_dashboard(wb, stats, keyword)
    buat_sheet_posts(wb, posts)
    buat_sheet_komentar_negatif(wb, komentar_negatif)
    buat_sheet_semua_komentar(wb, semua_komentar)

    wb.save(filename)
    print(f"✅ Excel tersimpan: {filename}")
    return filename
